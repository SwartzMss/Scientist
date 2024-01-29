import pandas as pd
from openpyxl import load_workbook
import os
from datetime import datetime

class ExcelProjectManagerUpdated:
    def __init__(self, project_name):
        self.project_name = project_name
        current_date = datetime.now()
        self.year = current_date.year
        self.month = current_date.month
        self.filename = f"aridropList_{self.year}_{str(self.month).zfill(2)}.xlsx"
        self.cached_data = []

    def start_service(self):
        # 检查文件是否存在，如果不存在则创建
        if not os.path.exists(self.filename):
            with pd.ExcelWriter(self.filename, engine='openpyxl') as writer:
                pass  # 仅创建文件

    def update_info(self, name, msg):
        # 缓存信息，稍后写入文件
        self.cached_data.append({"name": name, "msg": msg, "day": datetime.now().day})

    def save_msg_and_stop_service(self):
        # 将缓存的信息写入sheet
        try:
            book = load_workbook(self.filename)
            writer = pd.ExcelWriter(self.filename, engine='openpyxl')
            writer.book = book

            for data in self.cached_data:
                sheet_name = f"{self.project_name}_{str(self.month).zfill(2)}_{str(data['day']).zfill(2)}"
                if sheet_name in book.sheetnames:
                    # 如果工作表已存在，先删除旧的工作表
                    del book[sheet_name]

                # 创建新的工作表并写入数据
                df = pd.DataFrame([data])
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            writer.save()
        except Exception as e:
            print(f"Error while saving messages: {e}")

# 示例使用
# 创建类的实例
excel_manager_updated = ExcelProjectManagerUpdated("ProjectA")

# 启动服务，检查文件
excel_manager_updated.start_service()

# 更新信息
excel_manager_updated.update_info("Example Name", "Example Message")

# 保存消息并停止服务
excel_manager_updated.save_msg_and_stop_service()
